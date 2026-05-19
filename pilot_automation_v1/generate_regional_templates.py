import os
import pandas as pd
import openpyxl
import re
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

BASE_DIR = r"D:\ICBF\cost-tracking\pilot_automation_v1"
MASTER_FILE = os.path.join(BASE_DIR, "master_data", "ZonificacIón_Primera_Infancia_Hasta_Oct.xlsx")
TEMPLATE_FILE = os.path.join(BASE_DIR, "master_data", "MATRIZ SERVICIOS INTEGRALES 2026 REGIONAL BOLIVAR_IP (2).xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "shared_folders")

def clean_name(name):
    return str(name).strip().upper().replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U').replace('Ñ', 'N')

def update_formula(formula, source_row, target_row):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    
    delta = target_row - source_row
    
    def repl(m):
        prefix_col = m.group(1)
        col = m.group(2)
        prefix_row = m.group(3)
        row = int(m.group(4))
        
        # If the row is absolute (e.g. $A$5), don't change it
        if prefix_row:
            return m.group(0)
            
        return f"{prefix_col}{col}{prefix_row}{row + delta}"
        
    return re.sub(r'(\$?)([A-Z]+)(\$?)(\d+)', repl, formula)

def copy_cell(source_cell, target_cell, source_row_idx, target_row_idx):
    if source_cell.has_style:
        target_cell.font = openpyxl.styles.Font(**source_cell.font.__dict__)
        target_cell.border = openpyxl.styles.Border(**source_cell.border.__dict__)
        target_cell.fill = openpyxl.styles.PatternFill(**source_cell.fill.__dict__)
        target_cell.alignment = openpyxl.styles.Alignment(**source_cell.alignment.__dict__)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = openpyxl.styles.Protection(**source_cell.protection.__dict__)

    if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
        target_cell.value = update_formula(source_cell.value, source_row_idx, target_row_idx)
    else:
        target_cell.value = source_cell.value

def generate_templates():
    print("Starting Perfected Template Generation Process...", flush=True)
    xl_master = pd.ExcelFile(MASTER_FILE)
    
    MODALITY_CONFIG = {
        'Integrales': {
            'regional_idx': 1, 'cz_idx': 2, 'mun_idx': 3,
            'uds_code_idx': 4, 'uds_name_idx': 5,
            'modalidad_idx': 8, 'componente_idx': 9, 'cupos_idx': 10,
        },
        'Comunitarios': {
            'regional_idx': 0, 'cz_idx': 1, 'mun_idx': 2,
            'uds_code_idx': 4, 'uds_name_idx': 5,
            'modalidad_idx': 7, 'componente_idx': 6, 'cupos_idx': 10,
        }
    }
    
    for sheet_name, cfg in MODALITY_CONFIG.items():
        print(f"\n--- Modality: {sheet_name} ---", flush=True)
        df_master = pd.read_excel(xl_master, sheet_name=sheet_name, header=None)
        reg_col = cfg['regional_idx']
        df_data = df_master.dropna(subset=[reg_col])
        unique_regionals = [r for r in df_data[reg_col].unique() if "REGIONAL" not in str(r).upper() and "TIPO" not in str(r).upper()]
        
        for regional in unique_regionals:
            reg_clean = clean_name(regional)
            if reg_clean in ['', 'NAN']: continue
            
            output_path = os.path.join(OUTPUT_DIR, reg_clean, f"MATRIZ_2026_{reg_clean}_{sheet_name.upper()}.xlsx")
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            print(f"  > Processing {regional}...", flush=True)
            
            wb = openpyxl.load_workbook(TEMPLATE_FILE)
            ws = wb['Matriz']
            
            # Hide Matriz (2) sheet to keep the file clean and avoid corruption alerts
            if 'Matriz (2)' in wb.sheetnames:
                wb['Matriz (2)'].sheet_state = 'veryHidden'
            
            # Filter regional data
            df_reg = df_data[df_data[reg_col] == regional]
            num_rows_needed = len(df_reg)
            
            source_row_idx = 3
            source_cells = [ws.cell(row=source_row_idx, column=c) for c in range(1, ws.max_column + 1)]
            
            # Identify formula columns
            is_formula_col = {c: (isinstance(ws.cell(row=source_row_idx, column=c).value, str) and ws.cell(row=source_row_idx, column=c).value.startswith('=')) for c in range(1, ws.max_column + 1)}
            
            # Mapping
            col_mapping = {
                1: cfg['regional_idx'],
                2: cfg['cz_idx'],
                3: cfg['mun_idx'],
                4: cfg['componente_idx'],
                5: cfg['modalidad_idx'],
                6: cfg['uds_name_idx'],
                7: cfg['uds_code_idx'],
                8: cfg['cupos_idx']
            }
            
            for i, (idx, row_data) in enumerate(df_reg.iterrows(), start=3):
                if i > ws.max_row:
                    for col_idx, src_cell in enumerate(source_cells, start=1):
                        target_cell = ws.cell(row=i, column=col_idx)
                        copy_cell(src_cell, target_cell, source_row_idx, i)
                
                for col_idx in range(1, ws.max_column + 1):
                    if is_formula_col.get(col_idx, False):
                        continue # DO NOT TOUCH FORMULAS
                        
                    if col_idx in col_mapping:
                        val = row_data[col_mapping[col_idx]]
                        ws.cell(row=i, column=col_idx).value = val if pd.notna(val) else ""
                    else:
                        ws.cell(row=i, column=col_idx).value = ""
            
            # For unused rows at the bottom (if any), clear unmapped values to hide Bolivar leftovers
            for i in range(num_rows_needed + 3, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    if not is_formula_col.get(col_idx, False):
                        ws.cell(row=i, column=col_idx).value = ""
            
            # Do NOT expand or modify table ref (tbl.ref) to avoid "Removed Part: External data range" corruption alerts.
            # openpyxl cannot safely resize QueryTables/ExternalData ranges.
            
            wb.save(output_path)
            print(f"    Saved: {os.path.basename(output_path)}", flush=True)

if __name__ == "__main__":
    generate_templates()

