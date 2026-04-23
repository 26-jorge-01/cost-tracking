import pandas as pd
import glob
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- CONFIGURACIÓN DE RUTAS ---
DIR_BASE = r"D:\ICBF\cost-tracking\data\insumos metaverso"
ORIGINAL_METAVERSO = os.path.join(DIR_BASE, "Metaverso 2026.xlsx")
INTEGRALES_DIR = os.path.join(DIR_BASE, "Nivelacion de Canastas Integrales")
HCB_DIR = os.path.join(DIR_BASE, "Nivelación de Canastas HCB")
ALIMENTOS_PATH = os.path.join(DIR_BASE, "CONTRATOS ALIMENTOS ORGANIZACIONES CAMPESINAS.xlsx")

# El archivo final se guardará en la carpeta de insumos para mantener el orden del repositorio
OUTPUT_PATH = os.path.join(DIR_BASE, "RECONSTRUCCION_METAVERSO_2026_NACIONAL_AUDITADO.xlsx")

def find_sheet_robust(xl_names, search_terms):
    """Busca una hoja por términos clave de forma insensible a mayúsculas."""
    for name in xl_names:
        if all(term.upper() in name.upper() for term in search_terms):
            return name
    return None

def extract_all_insumos():
    """Recopila todos los datos frescos de los insumos 2026 (Nacional)."""
    print(">>> Extrayendo datos de Insumos 2026 (Integrales, HCB, Alimentos)...")
    rows = []
    
    # 1. INTEGRALES (Por Regional)
    int_files = glob.glob(os.path.join(INTEGRALES_DIR, "**", "*.xls*"), recursive=True)
    for f in int_files:
        if "~$" in f: continue
        try:
            xl = pd.ExcelFile(f)
            sh_zon = find_sheet_robust(xl.sheet_names, ['ZONIFICA', 'PEDAGOGICO'])
            if sh_zon:
                df = pd.read_excel(f, sheet_name=sh_zon, skiprows=4)
                for _, r in df.iterrows():
                    code = str(r.iloc[6]).split('.')[0]
                    if len(code) >= 10: # Validación básica de código UDS
                        rows.append({
                            'Codigo Unidad Servicio UDS': code,
                            'Cupos a Programar 2026': r.iloc[8],
                            'VALOR SOLO 2026': r.iloc[13],
                            'SERVICIO\n2026': r.iloc[4],
                            'ZONA 2026': str(r.iloc[0]),
                            'Modalidad 2026': 'INTEGRAL'
                        })
        except Exception as e: print(f"Error en {os.path.basename(f)}: {e}")

    # 2. HCB (Por Regional)
    hcb_files = glob.glob(os.path.join(HCB_DIR, "**", "*.xls*"), recursive=True)
    for f in hcb_files:
        if "~$" in f: continue
        try:
            xl = pd.ExcelFile(f)
            sh_hcb = find_sheet_robust(xl.sheet_names, ['BASE', 'COSTOS', 'HCB'])
            if sh_hcb:
                df = pd.read_excel(f, sheet_name=sh_hcb, skiprows=1)
                for _, r in df.iterrows():
                    code = str(r.iloc[6]).split('.')[0]
                    if len(code) >= 10:
                        rows.append({
                            'Codigo Unidad Servicio UDS': code,
                            'Cupos a Programar 2026': r.iloc[9],
                            'VALOR SOLO 2026': r.iloc[-1],
                            'Modalidad 2026': 'HCB'
                        })
        except Exception as e: print(f"Error en {os.path.basename(f)}: {e}")

    # 3. ALIMENTOS (Nacional)
    try:
        df_alim = pd.read_excel(ALIMENTOS_PATH, sheet_name='ZONIFICACION', skiprows=3)
        for _, r in df_alim.iterrows():
            code = str(r.iloc[7]).split('.')[0]
            if len(code) >= 10:
                rows.append({
                    'Codigo Unidad Servicio UDS': code,
                    'Cupos a Programar 2026': r.iloc[13],
                    'VALOR SOLO 2026': r.iloc[20],
                    'SERVICIO\n2026': r.iloc[9],
                    'CONCEPTO DEFINITIVO': r.iloc[14],
                    'NIT CONTRATISTA 2026': str(r.iloc[16]).split('.')[0],
                    'CONTRATISTA 2026': r.iloc[15],
                    'Modalidad 2026': 'ALIMENTOS'
                })
    except Exception as e: print(f"Error en Alimentos: {e}")
            
    return pd.DataFrame(rows).drop_duplicates(subset=['Codigo Unidad Servicio UDS', 'Modalidad 2026'])

def main_fusion_nacional():
    print(">>> Iniciando Reconstrucción Nacional (Fusión Histórico + Insumos)...")
    
    # Cargar Base Histórica
    df_old = pd.read_excel(ORIGINAL_METAVERSO, sheet_name='ZonificacionPais')
    df_old['Codigo Unidad Servicio UDS'] = df_old['Codigo Unidad Servicio UDS'].astype(str).str.split('.').str[0]
    headers_orig = df_old.columns.tolist()

    # Cargar Insumos Nuevos
    df_new = extract_all_insumos()
    
    # Fusión Maestra (Outer Join)
    df_final = pd.merge(df_old, df_new, on='Codigo Unidad Servicio UDS', how='outer', suffixes=('_old', ''))
    
    # Lógica de Auditoría
    old_codes = set(df_old['Codigo Unidad Servicio UDS'])
    new_codes = set(df_new['Codigo Unidad Servicio UDS'])
    
    def get_status(row):
        code = row['Codigo Unidad Servicio UDS']
        if code in old_codes and code in new_codes: return "ACTUALIZADO (INSUMO 2026)"
        if code in new_codes: return "NUEVO (SOLO INSUMO 2026)"
        return "HISTORICO (SIN RESPALDO 2026)"

    df_final['AUDITORIA_ESTADO_2026'] = df_final.apply(get_status, axis=1)
    
    # Guardar Excel
    final_cols = headers_orig + ['AUDITORIA_ESTADO_2026']
    df_final[final_cols].to_excel(OUTPUT_PATH, index=False)
    print(f"\n>>> Archivo Nacional generado: {OUTPUT_PATH}")
    
    # Aplicar Coloreado para Decisores
    print(">>> Aplicando formato visual de auditoría...")
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    fills = {
        'ACTUALIZADO': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'NUEVO': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        'HISTORICO': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    }
    
    col_idx = len(final_cols)
    for r in range(2, ws.max_row + 1):
        status = str(ws.cell(row=r, column=col_idx).value)
        for key, fill in fills.items():
            if key in status:
                for c in range(1, col_idx + 1):
                    ws.cell(row=r, column=c).fill = fill
                break
    wb.save(OUTPUT_PATH)
    print(">>> Proceso Nacional Finalizado.")

if __name__ == "__main__":
    main_fusion_nacional()
