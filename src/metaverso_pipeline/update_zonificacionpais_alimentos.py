import pandas as pd
import os
import unicodedata
import warnings
from datetime import datetime
warnings.filterwarnings('ignore')

DIR_BASE = r"D:\ICBF\cost-tracking\data\insumos metaverso"
INPUT_ZONIF = os.path.join(DIR_BASE, "ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx")
INPUT_ALIM  = os.path.join(DIR_BASE, "CONCAT_ALIMENTOS_2026.xlsx")
UDS_FILE    = os.path.join(DIR_BASE, "UDS_15052026_3112025.xlsx")
OUTPUT_ZONIF = os.path.join(DIR_BASE, "ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx")
OUTPUT_REPORTE = os.path.join(DIR_BASE, "REPORTE_NO_ACTUALIZABLES.xlsx")
OUTPUT_PERSISTENTES = os.path.join(DIR_BASE, "NO_ACTUALIZABLES_PERSISTENTES.xlsx")

def log(msg):
    print(f"[{datetime.now():%H:%M:%S}] {msg}")

def normalize(s):
    return unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('ASCII').lower().replace(' ', '').replace('\n', '')

def clean_nit(x):
    if pd.isna(x): return None
    s = str(x).replace(',', '.').split('.')[0]
    s = ''.join(c for c in s if c.isdigit())
    return s if s else None

print("="*60)
print("  ACTUALIZACION POR ALIMENTOS")
print("="*60)

# ============================================================
log("PASO 1: Cargar ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx")
df = pd.read_excel(INPUT_ZONIF, sheet_name='Sheet1')
ORIG_COLS = list(df.columns)
total_rows = len(df)
log(f"  {total_rows} filas, {len(ORIG_COLS)} columnas")

state_counts_before = df['Estado_Actualizacion'].value_counts().to_dict()
log(f"  Estados antes: {state_counts_before}")

# ============================================================
log("PASO 2: Cargar CONCAT_ALIMENTOS_2026.xlsx")
alim = pd.read_excel(INPUT_ALIM)
alim["k"] = alim["Codigo_UDS"].astype(str).str.strip()
alim_codes = set(alim["k"].unique())
log(f"  {len(alim)} filas, {len(alim_codes)} codigos unicos")

# ============================================================
log("PASO 3: Identificar registros No actualizable y Actualizado x integrales")
df["k_clean"] = df["Codigo Unidad Servicio UDS"].astype(str).str.strip().str.replace(".0", "", regex=False).replace(["nan", "None", "", " "], None)

target_mask = df['Estado_Actualizacion'].isin(["No actualizable", "Actualizado x integrales"])
target_count = target_mask.sum()
log(f"  Registros objetivo: 4991")
log(f"    No actualizable: {(df['Estado_Actualizacion'] == 'No actualizable').sum()}")
log(f"    Actualizado x integrales: {(df['Estado_Actualizacion'] == 'Actualizado x integrales').sum()}")

# ============================================================
log("PASO 4: Matching contra Alimentos")
target_codes = set(df.loc[target_mask, "k_clean"].dropna().unique())
overlap = target_codes & alim_codes
log(f"  Codigos objetivo unicos: {len(target_codes)}")
log(f"  Codigos en Alimentos: {len(overlap)}")

# Prepare Alimentos lookup (one row per UDS code, keep first)
alim_dedup = alim.drop_duplicates(subset="k", keep="first").set_index("k")

# ============================================================
log("PASO 5: Construir mapping de columnas")
MAPPING = [
    ("Servicio",                    "SERVICIO 2026"),
    ("Cupos",                       "Cupos a Programar 2026"),
    ("Componente_para_la_UDS",      "COMPONENTE"),
    ("CONTRATISTA",                 "CONTRATISTA 2026"),
    ("NIT_CONTRATISTA",             "NIT CONTRATISTA 2026"),
    ("ZONA",                        "ZONA 2026"),
    ("Regional_UDS",                "Regional UDS"),
    ("Municipio_UDS",               "Municipio UDS"),
    ("Centro_Zonal_UDS",            "Centro Zonal UDS"),
    ("UDS",                         "Unidad Servicio UDS"),
    ("CONCEPTO_DEFINITIVO",         "CONCEPTO DEFINITIVO"),
]

df_norm = {normalize(c): c for c in ORIG_COLS}
resolved = []
for src, dst in MAPPING:
    n = normalize(dst)
    if n in df_norm:
        resolved.append((src, df_norm[n]))
        log(f"  {src:38s} -> {df_norm[n].replace(chr(10), ' ')[:40]}")
    else:
        log(f"  {src:38s} -> NO ENCONTRADO en ZonificacionPais")

# Clean NIT values in Alimentos source
alim_dedup['NIT_CONTRATISTA'] = alim_dedup['NIT_CONTRATISTA'].apply(clean_nit)

# ============================================================
log("PASO 6: Aplicar actualizaciones")
matched_mask = target_mask & df['k_clean'].isin(overlap)
pre_count = matched_mask.sum()
log(f"  Filas a actualizar por Alimentos: {pre_count}")

changes_detail = []
for src, dst in resolved:
    change_count = 0
    matched_idx = df[matched_mask].index
    dst_dtype = df[dst].dtype
    for idx in matched_idx:
        ck = df.at[idx, 'k_clean']
        if ck in alim_dedup.index:
            new_val = alim_dedup.at[ck, src]
            if pd.notna(new_val):
                if isinstance(new_val, (int, float)) and not pd.isna(new_val):
                    if dst_dtype in (float, 'float64'):
                        new_val = float(new_val)
                    elif dst_dtype in (int, 'int64'):
                        new_val = int(new_val)
                    else:
                        new_val = str(new_val)
                elif not isinstance(new_val, str):
                    new_val = str(new_val)
                elif isinstance(new_val, str):
                    new_val = new_val.strip()
            old_val = df.at[idx, dst]
            if pd.notna(new_val):
                same = False
                try:
                    same = str(new_val) == str(old_val).strip() if pd.notna(old_val) else False
                except Exception:
                    same = False
                if not same:
                    changes_detail.append({
                        'k': ck,
                        'columna': dst,
                        'antes': old_val,
                        'despues': new_val,
                    })
                    change_count += 1
                    df.at[idx, dst] = new_val
    log(f"  {src.replace(chr(10), ' ')[:38]:38s}: {change_count} cambios de valor")

# Special handling: update Modalidad to ALIMENTOS
alim_mask = df['k_clean'].isin(overlap)
modalidad_cols = [c for c in ORIG_COLS if normalize(c) == normalize("Modalidad 2026")]
for mc in modalidad_cols:
    df.loc[alim_mask & target_mask, mc] = "ALIMENTOS"

# Set new status
df.loc[matched_mask, 'Estado_Actualizacion'] = "Actualizado x alimentos"
post_count = matched_mask.sum()
log(f"  Actualizados a 'Actualizado x alimentos': {post_count}")

# ============================================================
log("PASO 7: Identificar persistentes (siguen sin match)")
persistentes_mask = target_mask & ~df['k_clean'].isin(overlap)
persistentes_count = persistentes_mask.sum()
log(f"  Aun sin coincidencia (persistentes): {persistentes_count}")

# ============================================================
log("PASO 8: Generar REPORTE_NO_ACTUALIZABLES.xlsx (todos los registros objetivo)")
reporte_cols_base = [
    "Codigo Unidad Servicio UDS", "Unidad Servicio UDS",
    "Regional UDS", "Municipio UDS", "Departamento UDS",
    "COMPONENTE", "Modalidad 2026",
    "Cupos a Programar 2026", "Estado_Actualizacion",
]
reporte_cols = [c for c in reporte_cols_base if c in df.columns]
# Add service columns by flexible name matching
svc_cols = [c for c in ORIG_COLS if ('SERVICIO' in str(c).upper() and '2026' in str(c)) or c == 'Servicio 2025']
for c in svc_cols:
    if c not in reporte_cols:
        reporte_cols.append(c)

# Include all records that were ever in the "not actualizable" category:
# - No actualizable (was never matched)
# - Actualizado x integrales (was rescued by integrales, not alimentos)
# - Actualizado x alimentos (was rescued by this Alimentos step)
all_objective = df['Estado_Actualizacion'].isin(["No actualizable", "Actualizado x integrales", "Actualizado x alimentos"])
reporte = df[all_objective].copy()
reporte["Match_Alimentos"] = reporte['Estado_Actualizacion'] == "Actualizado x alimentos"
reporte["Razon_No_Actualizable"] = reporte.apply(
    lambda r: "ACTUALIZADO VIA ALIMENTOS" if r["Match_Alimentos"]
    else "SIN COINCIDENCIA EN ALIMENTOS (rescatado x integrales)" if r["Estado_Actualizacion"] == "Actualizado x integrales"
    else "SIN COINCIDENCIA EN ALIMENTOS",
    axis=1
)

reporte_out = reporte[reporte_cols + ["Match_Alimentos", "Razon_No_Actualizable"]] if reporte_cols else reporte
reporte_out.to_excel(OUTPUT_REPORTE, index=False)
log(f"  Reporte guardado: {OUTPUT_REPORTE} ({len(reporte_out)} filas)")

# ============================================================
log("PASO 9: Generar NO_ACTUALIZABLES_PERSISTENTES.xlsx (aun sin match)")
persistentes = df[persistentes_mask].copy()
persistentes["Razon"] = "Sin coincidencia en Alimentos ni en CONCAT (Comunitarios/Integrales)"
persistentes_out = persistentes[reporte_cols + ["Razon"]] if reporte_cols else persistentes
persistentes_out.to_excel(OUTPUT_PERSISTENTES, index=False)
log(f"  Persistentes guardado: {OUTPUT_PERSISTENTES} ({len(persistentes_out)} filas)")

# ============================================================
log("PASO 10: Guardar ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx actualizado")
df.drop(columns=['k_clean'], inplace=True, errors='ignore')
df.to_excel(OUTPUT_ZONIF, index=False, engine='openpyxl')
log(f"  OK: {OUTPUT_ZONIF} ({len(df)} filas x {len(df.columns)} cols)")

# ============================================================
log("PASO 11: Inyectar formulas en Unnamed")
import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook(OUTPUT_ZONIF)
ws = wb.active
final_cols = list(df.columns)
unnamed = None
riesgo_col = None
for c in final_cols:
    if str(c).startswith("Unnamed"):
        unnamed = c
    elif normalize(c) == normalize("RIESGO TRANSPARENCIA"):
        riesgo_col = c

if unnamed and riesgo_col:
    riesgo_idx = final_cols.index(riesgo_col)
    unnamed_idx = final_cols.index(unnamed)
    ref_letter = get_column_letter(riesgo_idx + 1)
    target_letter = get_column_letter(unnamed_idx + 1)
    nrows = len(df)
    batch = 5000
    for start in range(2, nrows + 2, batch):
        end = min(start + batch - 1, nrows + 1)
        for r in range(start, end + 1):
            ws[f"{target_letter}{r}"] = f"=+EXACT({ref_letter}{r},CF{r})"
    wb.save(OUTPUT_ZONIF)
    log(f"  {nrows} formulas inyectadas en {target_letter}2:{target_letter}{nrows+1}")

print()
print("="*60)
print("  RESUMEN DE ACTUALIZACION POR ALIMENTOS")
print("="*60)
print(f"  Archivo: {OUTPUT_ZONIF}")
print(f"  Total filas: {total_rows}")
print()
print("  Estado_Actualizacion ANTES:")
for s, n in sorted(state_counts_before.items()):
    print(f"    {s}: {n}")
print()
print("  Estado_Actualizacion DESPUES:")
for s, n in df['Estado_Actualizacion'].value_counts().items():
    print(f"    {s}: {n}")
print()
print("  Registros objetivo: 4991")
print(f"  Matched con Alimentos: {pre_count}")
print(f"  Persistentes sin match: {persistentes_count}")
print(f"  Cambios de valor: {len(changes_detail)}")
print()
print("  ARCHIVOS GENERADOS:")
print(f"    {OUTPUT_ZONIF}")
print(f"    {OUTPUT_REPORTE}")
print(f"    {OUTPUT_PERSISTENTES}")
print("="*60)
