import pandas as pd
import os
import unicodedata
import warnings
from datetime import datetime
warnings.filterwarnings('ignore')

DIR_BASE = r"D:\ICBF\cost-tracking\data\insumos metaverso"
CONCAT   = os.path.join(DIR_BASE, "CONCAT_ZONIFICACION_METAVERSO_2026.xlsx")
UDS      = os.path.join(DIR_BASE, "UDS_15052026_3112025.xlsx")
ORIG     = os.path.join(DIR_BASE, "Metaverso 2026.xlsx")
OUTPUT   = os.path.join(DIR_BASE, "ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx")

def log(msg):
    print(f"[{datetime.now():%H:%M:%S}] {msg}")

def normalize(s):
    return unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('ASCII').lower().replace(' ', '').replace('\n', '')

def clean_nit(x):
    if pd.isna(x): return None
    s = str(x).replace(',', '.').split('.')[0]
    s = ''.join(c for c in s if c.isdigit())
    return s if s else None

MONETARY = {
    " APALANCAMIENTO NUEVO DOS DIAS ",
    " VALOR TOTAL 2026 NUEVO ",
    " VALOR SOLO 2026 ",
    "APALANCA MAS VIGENCIA NUEVO 2026",
    "Costo total hasta Julio ajuste operador alimentos",
}

# ============================================================
log("PASO 1: Cargar ZonificacionPais original")
orig = pd.read_excel(ORIG, sheet_name="ZonificacionPais")
ORIG_COLS = list(orig.columns)
log(f"  {len(ORIG_COLS)} cols, {len(orig)} filas")

orig_norm = {normalize(c): c for c in ORIG_COLS if c not in MONETARY}

orig["k"] = orig["Codigo Unidad Servicio UDS"].astype(str).str.strip()
orig["k"] = orig["k"].replace(["nan", "None", ""], None)
orig_codes = set(orig.loc[orig["k"].notna(), "k"].unique())

# ============================================================
log("PASO 2: Construir datos nuevos")
df = pd.read_excel(CONCAT)
df["k"] = df["Codigo_UDS"].astype(str).str.strip()
log(f"  CONCAT: {len(df)} filas")

uds = pd.read_excel(UDS, sheet_name="UDS_15052026")
uds["k"] = uds["CodigoUnidadServicioUDS"].astype(str).str.strip()
uds = uds.drop_duplicates(subset="k", keep="first")
df = df.merge(uds, on="k", how="left", suffixes=("_cat", "_uds"))

th = pd.read_excel(UDS, sheet_name="TH_Transito_15052026")
th["k"] = th["codigounidadservicio"].astype(str).str.strip()
th = th.drop_duplicates(subset="k", keep="first")
df = df.merge(th[["k", "total"]].rename(columns={"total": "TH_Transito_Cupos"}), on="k", how="left")

uds25 = pd.read_excel(UDS, sheet_name="UDS_31122025")
uds25["k"] = uds25["CodigoUnidadServicioUDS"].astype(str).str.strip()
uds25 = uds25.drop_duplicates(subset="k", keep="first")
df = df.merge(uds25[["k", "CuposUDS_31122025", "CuposServicioUDS_31122025"]], on="k", how="left")

df["NIT_EC"] = df["NumeroDocumentoEC"].apply(clean_nit)
df["NumeroContrato_clean"] = df["NumeroContrato"].astype(str).str.strip()

AM = {"HCB":"FAMILIAR Y COMUNITARIA","FAMI":"FAMILIAR Y COMUNITARIA","HCB_SA":"FAMILIAR Y COMUNITARIA","BV":"INSTITUCIONAL","JC":"INSTITUCIONAL"}
def mod(row):
    if pd.notna(row["Abrev"]) and row["Abrev"] in AM: return AM[row["Abrev"]]
    s = str(row["Servicio"]).upper() if pd.notna(row["Servicio"]) else ""
    if "PROPIA" in s or "INTERCULTURAL" in s: return "PROPIA E INTERCULTURAL"
    if "INSTITUCIONAL" in s or "CDI" in s: return "INSTITUCIONAL"
    if "MODELO PROPIO" in s or "MAI" in s: return "MODELO PROPIO"
    if "FAMILIAR" in s or "COMUNITARIA" in s or "HCB" in s: return "FAMILIAR Y COMUNITARIA"
    return row["Abrev"] if pd.notna(row["Abrev"]) else None
df["Modalidad_calc"] = df.apply(mod, axis=1).fillna(df["Tipo_Modalidad"])
new_keys = set(df["k"].unique())
log(f"  {len(df)} filas, {len(new_keys)} codigos")

# ============================================================
log("PASO 3: Mapeo de columnas")
MAPPINGS = [
    ("Regional_UDS","Regional UDS"),("Centro_Zonal_UDS","Centro Zonal UDS"),
    ("Municipio_UDS","Municipio UDS"),("ZONA","ZONA 2026"),
    ("Codigo_UDS","Codigo Unidad Servicio UDS"),("UDS","Unidad Servicio UDS"),
    ("LITERAL_DE_CONTRATACION","TIPO DE CONTRATACION 2026"),
    ("Servicio","SERVICIO 2026"),("Cupos","Cupos a Programar 2026"),
    ("Madres_Unds","Cantidad Madres Comunitarias en la UDS"),
    ("Componente_para_la_UDS","COMPONENTE"),
    ("DepartamentoUDS","Departamento UDS"),("BarrioUDS","Barrio UDS"),
    ("DireccionUDS","Direccion UDS"),("EstadoUDS","Estado UDS"),
    ("CuposUDS","Cupos UDS ofertados 2025"),
    ("EntidadContratista","CONTRATISTA 2026"),("NIT_EC","NIT CONTRATISTA 2026"),
    ("NumeroContrato_clean","Numero Contrato"),
    ("TipoOrganizacionEC","Tipo de Organizacion"),
    ("DepartamentoEC","Departamento Entidad Contratista 2025"),
    ("MunicipioEC","Municipio Entidad Contratista 2025"),
    ("CodigoMunicipioEC","Codigo Municipio Entidad Contratista 2025"),
    ("ZonaUbicacionUDS","Zona Ubicacion UDS"),
    ("VigenciaServicio","Vigencia UDS"),("CodigoRegionalUDS","CodigoRegional"),
    ("NombrePropiedadInfraestructura","Propiedad de la infraestructura"),
    ("Modalidad_calc","Modalidad 2026"),
    ("EntidadContratista","NIT_EntidadContratista"),
    ("Servicio","Servicio 2025"),("Modalidad_calc","Modalidad 2025"),
]
resolved = []
for src, dst in MAPPINGS:
    n = normalize(dst)
    if n in orig_norm:
        resolved.append((src, orig_norm[n]))

# ============================================================
log("PASO 4: Merge vectorizado")
df_dedup = df.drop_duplicates(subset="k", keep="first").set_index("k")
up_cols = list({src for src,_ in resolved}) + ["CuposUDS_31122025","CuposServicioUDS_31122025","TH_Transito_Cupos"]
up_cols = [c for c in up_cols if c in df_dedup.columns]
dup = df_dedup[up_cols].rename(columns={c:f"_n_{c}" for c in up_cols})

merged = orig[["k"]].merge(dup, left_on="k", right_index=True, how="left")

audit = pd.Series("No actualizable", index=orig.index, dtype="string")
audit[merged["k"].isin(new_keys)] = "Actualizado"
audit[orig["k"].isna()] = "Sin codigo"
orig["Estado_Actualizacion"] = audit

for src, dst in resolved:
    nc = f"_n_{src}"
    if nc in merged.columns:
        m = merged[nc].notna()
        orig.loc[m.values, dst] = merged.loc[m.values, nc].values

NEW_COLS = ["CuposUDS_31122025","CuposServicioUDS_31122025","TH_Transito_Cupos"]
for c in NEW_COLS:
    nc = f"_n_{c}"
    orig[c] = orig.get(c) if nc not in merged.columns else merged[nc]

log(f"  Actualizados: {(audit=='Actualizado').sum()} | No act: {(audit=='No actualizable').sum()}")

# ============================================================
log("PASO 5: Nuevos UDS")
exist = set(orig.loc[orig["k"].notna(), "k"].unique())
nuevos = new_keys - exist
log(f"  {len(nuevos)} nuevos")

if nuevos:
    nr = df_dedup.loc[df_dedup.index.isin(nuevos)].copy()
    ndf = pd.DataFrame(index=range(len(nr)), columns=ORIG_COLS, dtype=object)
    for src, dst in resolved:
        if src in nr.columns:
            ndf[dst] = nr[src].values
    for c in NEW_COLS:
        if c in nr.columns:
            ndf[c] = nr[c].values
    ndf["Estado_Actualizacion"] = "Nuevo"
    orig = pd.concat([orig, ndf], ignore_index=True)

# ============================================================
log("PASO 6: Limpiar NITs")
for c in orig.columns:
    if "NIT_EntidadContratista" in str(c): continue
    ca = "".join(ch for ch in c if ch.isascii()).lower().strip()
    if "nit" in ca and "contratista" in ca:
        orig[c] = orig[c].apply(lambda x: str(int(x)) if pd.notna(x) and str(x).replace(".0","").isdigit() else (str(x) if pd.notna(x) else None))

# ============================================================
log("PASO 7: Guardar")
final_cols = [c for c in ORIG_COLS if c not in MONETARY and c != "k"]
for c in NEW_COLS:
    if c not in orig.columns: orig[c] = None

unnamed = None
for c in final_cols:
    if str(c).startswith("Unnamed"):
        unnamed = c; final_cols.remove(c); break

final_cols += [c for c in NEW_COLS if c not in final_cols] + ["Estado_Actualizacion"]
if unnamed: final_cols.append(unnamed)
final_cols = [c for c in final_cols if c in orig.columns]
out = orig[final_cols]

log(f"  Guardando con openpyxl...")
out.to_excel(OUTPUT, index=False, engine='openpyxl')
log(f"  OK: {OUTPUT} ({len(out)} filas x {len(final_cols)} cols)")

# ============================================================
log("PASO 8: Inyectar formulas en Unnamed: 101")
if unnamed:
    import openpyxl
    from openpyxl.utils import get_column_letter

    riesgo_col_name = None
    for c in final_cols:
        if normalize(c) == normalize("RIESGO TRANSPARENCIA"):
            riesgo_col_name = c
            break

    if riesgo_col_name:
        riesgo_idx = final_cols.index(riesgo_col_name)
        unnamed_idx = final_cols.index(unnamed)
        ref_letter = get_column_letter(riesgo_idx + 1)  # 1-indexed
        target_letter = get_column_letter(unnamed_idx + 1)

        log(f"  Riesgo: {riesgo_col_name} ({ref_letter}), Unnamed ({target_letter})")

        wb = openpyxl.load_workbook(OUTPUT)
        ws = wb.active
        nrows = len(out)
        batch = 5000
        for start in range(2, nrows + 2, batch):
            end = min(start + batch - 1, nrows + 1)
            for r in range(start, end + 1):
                ws[f"{target_letter}{r}"] = f"=+EXACT({ref_letter}{r},CF{r})"
        wb.save(OUTPUT)
        log(f"  {nrows} formulas inyectadas en {target_letter}2:{target_letter}{nrows+1}")
    else:
        log("  AVISO: No se encontro columna RIESGO TRANSPARENCIA")
elif any(str(c).startswith("Unnamed") for c in final_cols):
    log("  AVISO: Unnamed column exists but wasn't tracked")

# ============================================================
print("="*50)
print("RESUMEN:")
for s, n in orig["Estado_Actualizacion"].value_counts().items():
    print(f"  {s}: {n}")
print(f"  Filas: {len(out)}, Columnas: {len(final_cols)}")
th_codes = set(th["k"].unique())
print(f"  TH_Transito sin match: {len(th_codes-new_keys)}/{len(th_codes)}")
for c in NEW_COLS:
    if c in out.columns: print(f"  {c} vacios: {out[c].isna().sum()}/{len(out)}")
print()
