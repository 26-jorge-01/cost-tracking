import pandas as pd
import unicodedata
import warnings
from datetime import datetime
warnings.filterwarnings('ignore')

DIR_BASE = r"D:\ICBF\cost-tracking\data\insumos metaverso"
INPUT  = f"{DIR_BASE}\\ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx"
INTEGRALES = f"{DIR_BASE}\\consolidacion_matriz_integrales_28042026_COPIA_SEGURA.xlsx"
OUTPUT = f"{DIR_BASE}\\ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx"

SHEET_INT = "ZONIFICACI"+chr(211)+"N- PEDAGOGICO"

def log(msg):
    print(f"[{datetime.now():%H:%M:%S}] {msg}")

def norm(s):
    return unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('ASCII').lower().replace(' ', '').replace('\n', '')

print("=" * 60)
print("  ACTUALIZACION COMPLEMENTARIA: Integrales Zonificacion")
print("=" * 60)

log("PASO 1: Cargar ZONIFICACIONPAIS_ACTUALIZADA_2026.xlsx")
df = pd.read_excel(INPUT)
ORIG_COLS = list(df.columns)
total_rows = len(df)
log(f"  {total_rows} filas, {len(ORIG_COLS)} columnas")

log(f"\nPASO 2: Cargar Integrales ZONIFICACION ({SHEET_INT})")
int_z = pd.read_excel(INTEGRALES, sheet_name=SHEET_INT)
log(f"  {len(int_z)} filas")

log("PASO 3: Identificar registros No actualizable")
k = df['Codigo Unidad Servicio UDS'].astype(str).str.strip()
k = k.replace(['nan', 'None', ''], None)
no_act_mask = df['Estado_Actualizacion'] == 'No actualizable'
no_act_count = no_act_mask.sum()
log(f"  No actualizable: {no_act_count}")

log("PASO 4: Matching de codigos UDS")
int_z['k'] = int_z['Codigo Unidad Servicio UDS'].apply(
    lambda x: str(int(float(str(x)))) if pd.notna(x) else None
)
int_lookup = int_z.drop_duplicates(subset='k', keep='first').set_index('k')
int_codes = set(int_lookup.index)
log(f"  Integrales ZONIFICACION codigos unicos: {len(int_codes)}")

df['k_clean'] = df.apply(
    lambda r: str(r['Codigo Unidad Servicio UDS']).replace('.0', '')
    if pd.notna(r['Codigo Unidad Servicio UDS']) else None,
    axis=1
)
no_act_codes = set(df.loc[no_act_mask, 'k_clean'].dropna().unique())
log(f"  No actualizable codigos unicos: {len(no_act_codes)}")

overlap = no_act_codes & int_codes
log(f"  Codigos a actualizar desde Integrales: {len(overlap)}")

if not overlap:
    log("  No hay codigos para actualizar. Saliendo...")
    print("\n" + "=" * 60)
    print("RESUMEN:")
    print(f"  No actualizable originales: {no_act_count}")
    print(f"  Actualizados desde Integrales: 0")
    print("=" * 60)
    exit()

log(f"\nPASO 5: Construir mapping de columnas Integrales -> ZonificacionPais")
MAPPING = [
    ("Regional UDS", "Regional UDS"),
    ("Centro Zonal UDS", "Centro Zonal UDS"),
    ("Municipio UDS", "Municipio UDS"),
    ("SERVICIO 2026", "SERVICIO 2026"),
    ("Unidad Servicio UDS", "Unidad Servicio UDS"),
]

df_norm = {norm(c): c for c in ORIG_COLS}
resolved = []
for src, dst in MAPPING:
    n = norm(dst)
    if n in df_norm:
        resolved.append((src, df_norm[n]))
        log(f"  {src:40s} -> {df_norm[n]}")
    else:
        log(f"  {src:40s} -> NO ENCONTRADO en ZonificacionPais")

update_cols = [dst for _, dst in resolved]
log(f"  Columnas a actualizar: {len(update_cols)}")

log("\nPASO 6: Actualizar registros")
matched_mask = no_act_mask & df['k_clean'].isin(overlap)
pre_count = matched_mask.sum()
log(f"  Filas a modificar: {pre_count}")

changes_detail = []
for src, dst in resolved:
    change_count = 0
    matched_idx = df[matched_mask].index
    for idx in matched_idx:
        ck = df.at[idx, 'k_clean']
        if ck in int_lookup.index:
            new_val = int_lookup.at[ck, src]
            old_val = df.at[idx, dst]
            if pd.notna(new_val) and str(new_val).strip() != str(old_val).strip():
                changes_detail.append({
                    'k': ck,
                    'columna': dst,
                    'antes': old_val,
                    'despues': new_val,
                })
                change_count += 1
            df.at[idx, dst] = new_val
    log(f"  {src:40s}: {change_count} cambios de valor")

df.loc[matched_mask, 'Estado_Actualizacion'] = 'Actualizado x integrales'
post_count = matched_mask.sum()

log(f"\nPASO 7: Guardar")
df.drop(columns=['k_clean'], inplace=True)
df.to_excel(OUTPUT, index=False, engine='openpyxl')
log(f"  Guardado: {OUTPUT} ({len(df)} filas x {len(df.columns)} cols)")

log("\nPASO 8: Inyectar formulas en Unnamed: 101")
import openpyxl
from openpyxl.utils import get_column_letter

final_cols = list(df.columns)
unnamed = None
riesgo_col = None
for c in final_cols:
    if str(c).startswith("Unnamed"):
        unnamed = c
    elif norm(c) == norm("RIESGO TRANSPARENCIA"):
        riesgo_col = c

if unnamed and riesgo_col:
    riesgo_idx = final_cols.index(riesgo_col)
    unnamed_idx = final_cols.index(unnamed)
    ref_letter = get_column_letter(riesgo_idx + 1)
    target_letter = get_column_letter(unnamed_idx + 1)

    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb.active
    nrows = len(df)
    batch = 5000
    for start in range(2, nrows + 2, batch):
        end = min(start + batch - 1, nrows + 1)
        for r in range(start, end + 1):
            ws[f"{target_letter}{r}"] = f"=+EXACT({ref_letter}{r},CF{r})"
    wb.save(OUTPUT)
    log(f"  {nrows} formulas inyectadas en {target_letter}2:{target_letter}{nrows+1}")

print("\n" + "=" * 60)
print("  RESUMEN DE ACTUALIZACION COMPLEMENTARIA")
print("=" * 60)
print(f"  Archivo: {OUTPUT}")
print(f"  Total filas: {total_rows}")
print(f"  No actualizable originales: {no_act_count}")
print(f"  Codigos Integrales ZONIFICACION unicos: {len(int_codes)}")
print(f"  Codigos en 'No actualizable': {len(no_act_codes)}")
print(f"  Codigos coincidentes (a actualizar): {len(overlap)}")
print(f"  Filas actualizadas: {pre_count}")
print(f"  Cambios de valor concretos: {len(changes_detail)}")
print()
print("  Nuevo Estado_Actualizacion:")
for v, c in df['Estado_Actualizacion'].value_counts().items():
    print(f"    {v}: {c}")
print()
print(f"  Columnas actualizadas: {update_cols}")
print()

if changes_detail:
    print("  Muestra de cambios (primeros 10):")
    print(f"    {'Codigo UDS':<20} {'Columna':<30} {'Antes':<30} {'Despues':<30}")
    print(f"    {'-'*110}")
    for ch in changes_detail[:10]:
        print(f"    {ch['k']:<20} {ch['columna']:<30} {str(ch['antes'])[:28]:<30} {str(ch['despues'])[:28]:<30}")
    if len(changes_detail) > 10:
        print(f"    ... y {len(changes_detail) - 10} mas")
print("=" * 60)
