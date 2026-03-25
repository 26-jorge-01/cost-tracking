# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd
from tqdm import tqdm
import warnings
from datetime import datetime
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed

# Importar configuración
try:
    from . import config
except ImportError:
    import config

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def encontrar_archivos_xlsm(ruta):
    ruta_path = Path(ruta)
    archivos_xlsm = list(ruta_path.rglob('*.xlsm'))
    return [str(p) for p in archivos_xlsm]

def contar_pdfs_en_ruta(ruta):
    ruta_path = Path(ruta)
    try:
        return len([archivo for archivo in ruta_path.iterdir() if archivo.is_file() and archivo.suffix == '.pdf'])
    except Exception: return 0

def listar_pdfs_en_ruta(ruta):
    ruta_path = Path(ruta)
    try:
        return [archivo.name for archivo in ruta_path.iterdir() if archivo.is_file() and archivo.suffix == '.pdf']
    except Exception: return []

def extraer_mes_del_nombre(nombre_archivo):
    nombre_archivo_lower = nombre_archivo.lower()
    for mes_key in config.MESES_MAP.keys():
        if mes_key in nombre_archivo_lower: return mes_key.capitalize()
    return 'Mes no encontrado'

def follow_formula(wb_raw, sheet_name, cell_ref, depth=0):
    if depth > 10: return None
    try:
        if sheet_name not in wb_raw.sheetnames: return None
        ws = wb_raw[sheet_name]
        val = ws[cell_ref].value
        if val is None: return None
        
        s_val = str(val).strip()
        if s_val.startswith('='):
            clean = s_val.replace('=', '').replace('+', '').replace('$', '').strip()
            if '!' in clean:
                parts = clean.split('!')
                target_s = parts[0].replace("'", "").strip()
                target_c = parts[1].strip()
                return follow_formula(wb_raw, target_s, target_c, depth + 1)
        return val
    except: return None

def procesar_archivo_xlsm(file_path_str, cells_to_extract, current_date, debug=False):
    file_path = Path(file_path_str)
    ruta_directorio = file_path.parent
    ultima_subcarpeta = ruta_directorio.name
    extracted_celdas, extracted_pdfs = [], []
    
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        wb_raw = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        
        # Identificar hojas de meses disponibles (Dinámico)
        month_names = [m.upper() for m in config.MESES_MAP.keys()]
        valid_sheets = [s for s in wb_data.sheetnames if s.upper() in month_names]
        
        if not valid_sheets:
            valid_sheets = [wb_data.sheetnames[0]]

        if debug:
            print(f"\n[DEBUG] Archivo: {file_path.name}")
            print(f"[DEBUG] Hojas detectadas: {valid_sheets}")
        
        num_pdfs = contar_pdfs_en_ruta(ruta_directorio)
        
        for sname in valid_sheets:
            row_celdas = [str(file_path), current_date, num_pdfs, sname]
            for cell_ref in cells_to_extract:
                val = wb_data[sname][cell_ref].value
                if val is None:
                    val = follow_formula(wb_raw, sname, cell_ref)
                row_celdas.append(val)
            extracted_celdas.append(row_celdas)
            
            # Llave para vinculación
            llave_val = wb_data[sname]['W29'].value
            if llave_val is None: llave_val = follow_formula(wb_raw, sname, 'W29')
            if not llave_val: llave_val = file_path.parent.name
            
            # Listar PDFs solo en la última hoja procesada para evitar redundancia en el detalle de PDFs
            if sname == valid_sheets[-1]:
                pdfs = listar_pdfs_en_ruta(ruta_directorio)
                for p in (pdfs if pdfs else ['No PDF']):
                    extracted_pdfs.append([
                        str(file_path), current_date, ultima_subcarpeta, 
                        p, extraer_mes_del_nombre(p) if p != 'No PDF' else 'N/A', 
                        llave_val, sname
                    ])
        
        wb_data.close()
        wb_raw.close()
    except Exception as e:
        if debug: print(f"[DEBUG] Error: {e}")
        extracted_celdas.append([str(file_path), current_date, 0, f"Error: {str(e)}"] + [None] * len(cells_to_extract))
        extracted_pdfs.append([str(file_path), current_date, ultima_subcarpeta, 'Error', 'N/A', 'Error', 'Error'])
        
    return extracted_celdas, extracted_pdfs

def main():
    c = config
    lista_xlsm = encontrar_archivos_xlsm(c.BASE_PATH)
    all_celdas, all_pdfs = [], []
    current_date = datetime.now().strftime("%d/%m/%Y")

    print(f"--- FASE 1: DESCUBRIMIENTO UNIVERSAL ---")
    if not lista_xlsm:
        print("No se encontraron archivos .xlsm.")
        return

    # Proceso del primero (Diagnóstico)
    res_c, res_p = procesar_archivo_xlsm(lista_xlsm[0], c.CELLS_TO_EXTRACT, current_date, debug=True)
    all_celdas.extend(res_c)
    all_pdfs.extend(res_p)

    restando = lista_xlsm[1:]
    if restando:
        with ProcessPoolExecutor() as executor:
            futures = {executor.submit(procesar_archivo_xlsm, f, c.CELLS_TO_EXTRACT, current_date, False): f for f in restando}
            for future in tqdm(as_completed(futures), total=len(futures), desc="Procesando resto"):
                try:
                    rc, rp = future.result()
                    all_celdas.extend(rc)
                    all_pdfs.extend(rp)
                except Exception as e: print(f"Error: {e}")

    # Organizar columnas
    col_names = [c.MAP_CELDAS.get(cell, f"Dato_{cell}") for cell in c.CELLS_TO_EXTRACT]
    final_cols = ['Ruta Archivo', 'Fecha Proceso', 'CP PDFs', 'Hoja Usada'] + col_names
    
    df_celdas = pd.DataFrame(all_celdas, columns=final_cols)
    
    # Manejo de Alias (duplicar columnas según configuración)
    if hasattr(c, 'COL_ALIASES'):
        for alias, source in c.COL_ALIASES.items():
            if source in df_celdas.columns:
                df_celdas[alias] = df_celdas[source]

    if c.COL_LLAVE in df_celdas.columns:
        df_celdas[c.COL_LLAVE] = df_celdas[c.COL_LLAVE].fillna(df_celdas['Ruta Archivo'].apply(lambda x: Path(x).parent.name))

    df_pdfs = pd.DataFrame(all_pdfs, columns=['Ruta Archivo', 'Fecha Proceso', 'Subcarpeta', 'Nombre PDF', 'Mes Detectado', 'Valor Validacion', 'Hoja Usada'])

    df_celdas.to_excel(c.BASE_PATH / c.OUTPUT_VALORES_CELDAS, index=False)
    df_pdfs.to_excel(c.BASE_PATH / c.OUTPUT_DETALLE_PDFS, index=False)
    print("\nFASE 1 COMPLETADA.")

if __name__ == '__main__':
    main()