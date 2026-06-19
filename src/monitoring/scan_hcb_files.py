import os, re, time, unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REGIONALES_DIR = Path(r"d:\ICBF\cost-tracking\data\replicacion hcb 5 junio\REGIONALES")
OUTPUT_DIR = Path(r"d:\ICBF\cost-tracking\data\monitoring")
OUTPUT_FILE = OUTPUT_DIR / "monitor_hcb.xlsx"
SHEET_NAME = "MATRIZ"

KEYWORDS = {
    "REGIONAL": ["REGIONAL"],
    "REF_CONTRATO": ["REFERENCIA", "CONTRATO SECOP"],
    "RP": ["No. RP"],
    "MUNICIPIO": ["MUNICIPIO"],
    "SERVICIO": ["NOMBRE DEL SERVICIO"],
    "CUPOS": ["SUMATORIA DE LOS CUPOS"],
    "EAS": ["NOMBRE EAS"],
    "NIT": ["NIT"],
    "VALOR_INI": ["VALOR TOTAL INICIAL APORTE ICBF"],
    "VALOR_ADC": ["VALOR ADICION NIVELACION CANASTA (UNICO"],
    "ALERTA_1000": ["ALERTA DE AVAL", "1000 SMLMV"],
    "ALERTA_5000": ["ALERTA CONTRATO DE APORTE", "5000 SMLMV"],
}

DICT = {
    "REGIONAL": "Nombre de la regional",
    "ARCHIVO": "Ruta del archivo MATRIZ",
    "FECHA_SCAN": "Momento del escaneo",
    "TAMANO_KB": "Tama\u00f1o del archivo en KB",
    "ULTIMA_MODIFICACION": "Ultima modificacion del archivo",
    "DIAS_DESDE_MOD": "Dias desde la ultima modificacion",
    "FILAS_DATOS": "Filas con datos en la hoja MATRIZ",
    "COLUMNAS": "Columnas en la hoja MATRIZ",
    "HOJAS": "Hojas del libro",
    "REGIONAL_EN_DATOS": "Regional en la primera fila de datos",
    "ERRORES_REF": "Celdas con error #REF!",
    "VACIOS_CRITICOS": "Celdas vacias en columnas clave",
    "PCT_VACIOS": "% celdas criticas vacias",
    "TIENE_ERRORES": "SI = tiene #REF!",
    "TIENE_VACIOS": "SI = tiene vacios criticos",
    "TIENE_ALERTAS": "SI = tiene alertas",
    "CONTRATOS_UNICOS": "Contratos distintos",
    "MUNICIPIOS": "Municipios distintos",
    "EAS_UNICOS": "Contratistas distintos",
    "SERVICIOS": "Tipos de servicio",
    "CUPOS_TOTALES": "Suma total de cupos",
    "CUPOS_X_CONTRATO": "Promedio cupos por contrato",
    "VALOR_INICIAL_TOTAL": "Valor inicial total",
    "VALOR_INICIAL_X_CONTRATO": "Valor inicial promedio",
    "VALOR_ADICIONAR_TOTAL": "Valor adicion total",
    "VALOR_ADICION_X_CONTRATO": "Valor adicion promedio",
    "ALERTA_1000": "Alertas >1000 SMLMV",
    "ALERTA_5000": "Alertas >5000 SMLMV",
    "ESTADO": "VERDE/AMARILLO/NARANJA/ROJO",
}


def _strip_accents(text):
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _match_col(headers, keywords):
    for col_idx, hdr in sorted(headers.items()):
        hdr_clean = _strip_accents(hdr.split("\n")[0]).strip().upper()
        for kw in keywords:
            if _strip_accents(kw).upper() in hdr_clean:
                return col_idx
    return None


def scan_file(filepath, folder_name):
    stats = filepath.stat()
    metrics = {
        "REGIONAL": folder_name,
        "ARCHIVO": str(filepath.resolve()),
        "FECHA_SCAN": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "TAMANO_KB": round(stats.st_size / 1024, 1),
        "ULTIMA_MODIFICACION": datetime.fromtimestamp(stats.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
    }

    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    metrics["COLUMNAS"] = ws.max_column or 70
    metrics["HOJAS"] = "|".join(wb.sheetnames)

    headers = {}
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        for ci, val in enumerate(row, start=1):
            if val:
                headers[ci] = str(val).strip()

    col_map = {k: _match_col(headers, kw) for k, kw in KEYWORDS.items()}

    data_rows = errors_ref = empty_critical = 0
    total_cupos = total_valor_ini = total_valor_adc = 0.0
    contracts, municipalities, eas_set, services = set(), set(), set(), set()
    alerta_1000 = alerta_5000 = 0
    regional_en_datos = ""
    crit_cols = [col_map[k] for k in ["REF_CONTRATO","MUNICIPIO","EAS","RP","SERVICIO"] if col_map.get(k)]
    first = True

    for row in ws.iter_rows(min_row=3, values_only=True):
        rd = {i+1: v for i, v in enumerate(row)}
        if all(v is None for v in rd.values()):
            continue
        data_rows += 1
        for v in rd.values():
            if isinstance(v, str) and "#REF!" in v:
                errors_ref += 1
        c = col_map
        if first and c.get("REGIONAL") and rd.get(c["REGIONAL"]):
            regional_en_datos = str(rd[c["REGIONAL"]]).strip()
            first = False
        if c.get("REF_CONTRATO") and rd.get(c["REF_CONTRATO"]):
            v = rd[c["REF_CONTRATO"]]
            contracts.add(str(int(v)) if isinstance(v, (int, float)) else str(v).strip())
        if c.get("MUNICIPIO") and rd.get(c["MUNICIPIO"]):
            municipalities.add(str(rd[c["MUNICIPIO"]]).strip().upper())
        if c.get("EAS") and rd.get(c["EAS"]):
            eas_set.add(str(rd[c["EAS"]]).strip().upper())
        if c.get("SERVICIO") and rd.get(c["SERVICIO"]):
            services.add(str(rd[c["SERVICIO"]]).strip())
        if c.get("CUPOS") and isinstance(rd.get(c["CUPOS"]), (int, float)):
            total_cupos += rd[c["CUPOS"]]
        if c.get("VALOR_INI") and isinstance(rd.get(c["VALOR_INI"]), (int, float)):
            total_valor_ini += rd[c["VALOR_INI"]]
        if c.get("VALOR_ADC") and isinstance(rd.get(c["VALOR_ADC"]), (int, float)):
            total_valor_adc += rd[c["VALOR_ADC"]]
        if c.get("ALERTA_1000") and rd.get(c["ALERTA_1000"]):
            v = str(rd[c["ALERTA_1000"]]).strip().upper()
            if v in ("SI","REQUIERE AVAL","ALERTA","ALERTA MAYOR QUE 1.000"):
                alerta_1000 += 1
        if c.get("ALERTA_5000") and rd.get(c["ALERTA_5000"]):
            v = str(rd[c["ALERTA_5000"]]).strip().upper()
            if v in ("SI","REQUIERE AVAL","ALERTA","ALERTA MAYOR QUE 5.000"):
                alerta_5000 += 1
        for cc in crit_cols:
            if rd.get(cc) is None:
                empty_critical += 1

    wb.close()
    metrics["REGIONAL_EN_DATOS"] = regional_en_datos
    metrics["FILAS_DATOS"] = data_rows
    metrics["ERRORES_REF"] = errors_ref
    metrics["VACIOS_CRITICOS"] = empty_critical
    metrics["CONTRATOS_UNICOS"] = len(contracts)
    metrics["MUNICIPIOS"] = len(municipalities)
    metrics["EAS_UNICOS"] = len(eas_set)
    metrics["SERVICIOS"] = "|".join(sorted(services)) if services else ""
    metrics["CUPOS_TOTALES"] = total_cupos
    metrics["VALOR_INICIAL_TOTAL"] = total_valor_ini
    metrics["VALOR_ADICIONAR_TOTAL"] = total_valor_adc
    metrics["ALERTA_1000"] = alerta_1000
    metrics["ALERTA_5000"] = alerta_5000
    return metrics


def add_calculated_fields(df):
    expected = df["FILAS_DATOS"] * 5
    df["PCT_VACIOS"] = (df["VACIOS_CRITICOS"] / expected.replace(0, 1) * 100).round(1)
    df["TIENE_ERRORES"] = df["ERRORES_REF"].apply(lambda x: "SI" if x > 0 else "NO")
    df["TIENE_VACIOS"] = df["VACIOS_CRITICOS"].apply(lambda x: "SI" if x > 0 else "NO")
    df["TIENE_ALERTAS"] = df.apply(lambda r: "SI" if r["ALERTA_1000"] + r["ALERTA_5000"] > 0 else "NO", axis=1)
    nz = df["CONTRATOS_UNICOS"].replace(0, 1)
    df["CUPOS_X_CONTRATO"] = (df["CUPOS_TOTALES"] / nz).round(0).astype("Int64")
    df["VALOR_INICIAL_X_CONTRATO"] = (df["VALOR_INICIAL_TOTAL"] / nz).round(0).astype("Int64")
    df["VALOR_ADICION_X_CONTRATO"] = (df["VALOR_ADICIONAR_TOTAL"] / nz).round(0).astype("Int64")
    now = datetime.now()
    df["DIAS_DESDE_MOD"] = df["ULTIMA_MODIFICACION"].apply(
        lambda x: (now - datetime.strptime(x, "%Y-%m-%d %H:%M:%S")).days if isinstance(x, str) else 0
    )

    def estado(r):
        if r["ERRORES_REF"] > 0: return "ROJO"
        if r["ALERTA_1000"] + r["ALERTA_5000"] > 0: return "NARANJA"
        if r["VACIOS_CRITICOS"] > 0: return "AMARILLO"
        return "VERDE"
    df["ESTADO"] = df.apply(estado, axis=1)
    return df


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    rows = []
    for d in sorted(REGIONALES_DIR.iterdir()):
        if not d.is_dir() or d.name.startswith("_"):
            continue
        fp = d / f"MATRIZ_{d.name}.xlsx"
        if not fp.exists():
            continue
        print(f"[SCAN] {d.name}...", end=" ", flush=True)
        t0 = time.time()
        try:
            row = scan_file(fp, d.name)
            rows.append(row)
            print(f"OK ({time.time()-t0:.1f}s) | {row['FILAS_DATOS']} filas")
        except Exception as e:
            print(f"ERROR: {e}")
    if not rows:
        print("No se encontraron archivos.")
        return
    df = add_calculated_fields(pd.DataFrame(rows))
    cols = ["REGIONAL","ARCHIVO","FECHA_SCAN","TAMANO_KB","ULTIMA_MODIFICACION","DIAS_DESDE_MOD",
            "FILAS_DATOS","COLUMNAS","HOJAS","REGIONAL_EN_DATOS",
            "ERRORES_REF","VACIOS_CRITICOS","PCT_VACIOS","TIENE_ERRORES","TIENE_VACIOS",
            "CONTRATOS_UNICOS","MUNICIPIOS","EAS_UNICOS","SERVICIOS",
            "CUPOS_TOTALES","CUPOS_X_CONTRATO","VALOR_INICIAL_TOTAL","VALOR_INICIAL_X_CONTRATO",
            "VALOR_ADICIONAR_TOTAL","VALOR_ADICION_X_CONTRATO",
            "ALERTA_1000","ALERTA_5000","TIENE_ALERTAS","ESTADO"]
    df = df[[c for c in cols if c in df.columns]]
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="DATOS", index=False)
        dict_df = pd.DataFrame([{"COLUMNA": k, "DESCRIPCION": v} for k, v in DICT.items()])
        dict_df.to_excel(writer, sheet_name="DICCIONARIO", index=False)

    print(f"\n=== COMPLETADO ===")
    print(f"  Archivos: {len(df)}")
    print(f"  Filas totales: {df['FILAS_DATOS'].sum():,}")
    for k, v in df['ESTADO'].value_counts().items():
        print(f"  {k}: {v}")
    print(f"  Salida: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
